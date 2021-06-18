import openpyxl
from openpyxl import Workbook
import platform
import os
from pathlib import Path
from src.handlers.dateformat_handler import *


def read_xlsx(file, date_format, selections, filelocation=None):
    error = None
    rows = []
    filename = ""
    try:
        filepath = Path(file)
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        title = sheet.title
        for idx in selections.curselection():
            for i, row in enumerate(sheet.rows):
                if i == 0:
                    rows.append([i.value for i in row])
                    continue
                row[idx].value = date_formatter(row[idx].value, date_format)
                rows.append([i.value for i in row])
        if filelocation is None:
            if platform.system() != 'Windows':
                error = 'Only Windows supported'
            else:
                username = os.getlogin()
                filename = 'C:\\Users\\%s\\Desktop\\%s' \
                           % (username, os.path.basename(file))
        else:
            error = 'Other locations not supported'
        new_wb = Workbook()
        ws = new_wb.worksheets[0]
        ws.title = title
        for row in rows:
            ws.append(row)
        wb.save(filename=filename)
    except Exception as e:
        print(e.with_traceback())
        error = 'Error: %s' % e
    return filename, error


def get_xlsx_headers(file):
    filepath = Path(file)
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active
    headers = list(sheet.iter_rows())[0]
    return [i.value for i in headers]
